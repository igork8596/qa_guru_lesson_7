from openpyxl import load_workbook
from os import path as p


def test_xlsx(way_to_dir):
    way_to_xlsx_file = p.join(way_to_dir, './resources/file_example_XLSX_50.xlsx')
    workbook = load_workbook(way_to_xlsx_file)
    sheet = workbook.active

    assert sheet.cell(row=3, column=2).value == 'Mara'
